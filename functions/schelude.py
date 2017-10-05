import openpyxl as xl
import os
import configparser
from speak import speak 
config = configparser.ConfigParser()
schldspath = 'data/schdls/'
#load confugiration file
config.read("data/main.cfg")


def getsch():
	return shldspath

def getsheet(path,degree,group):
	dirt = os.listdir("data/schdls")
	if degree+".xlsx" in dirt:
		wp = xl.load_workbook(path + degree + '.xlsx')
		sheets = wp.get_sheet_names()
		if group in sheets:
			return wp.get_sheet_by_name(group)

def getTC(day,sheet,tc):
	if tc in ['TD','Cours']:
		days = list()
		modules = list()
		module = list()
		for i in range(2,sheet.max_column+1):
			days.append(sheet.cell(row=1,column=i).value)
		for i in days:
			if i == day:
				for j in range(2,sheet.max_row+1):
						modules.append([sheet.cell(row=j,column=1).value,sheet.cell(row=j,column=days.index(day)+2).value])
		for i in modules:
			s = i[1].split(' ')
			#type TD or Cours
			if s[0]==tc:
				module.append(i)
		return module


def getAll(day,sheet):
	days = list()
	modules = list()
	cours = list()
	for i in range(2,sheet.max_column+1):
		days.append(sheet.cell(row=1,column=i).value)
	for i in days:
		if i == day:
			for j in range(2,sheet.max_row+1):
					modules.append([sheet.cell(row=j,column=1).value,sheet.cell(row=j,column=days.index(day)+2).value])
	return modules

def getTimeofmodule(module,sheet):
	days = getdays(sheet)
	for i in days:
		tds = getTds(i,sheet)
		for j in tds:
			if td in j[1]:
				return j[0]

def getdays(sheet):
	days = list()
	for i in range(2,sheet.max_column+1):
		days.append(sheet.cell(row=1,column=i).value)
	return days

def getModulebyTime(day,hour):
	group = ''
	degree = ''
	if 'user' in config:
		group = config['user']['group']
		degree = config['user']['degree']
	sheet = getsheet(schldspath,degree,group)
	days =  getdays(sheet)
	modules = getAll(day,sheet)
	for i in modules:
		time = i[0].split(' ')
		if hour == time[0]:
			return i[1]


def tellmodules(day,tc):
	#student list the value of group in index 7 and the degree in the 6 index
	group = ''
	degree = ''
	if 'user' in config:
		group = config['user']['group']
		degree = config['user']['degree']
	temp = list()
	if tc in ['TD','Cours']:
		tds = getTC(day,getsheet(schldspath,degree,group),tc)
		if len(tds) % 2 !=0:
			tds.append(['0','empty'])
		for i in tds:
			time = i[0]
			td = i[1]
			if temp == []:
				temp.append(time)
				temp.append(td)
			elif temp[1] == td:
					t1 = temp[0].split(' ')
					time = time.split(' ')
					speak("you have "+ td +" from "+ t1[0] +" to " + time[len(time)-1])
					temp.clear()
			elif time[0]=='0':
				time = temp[0].split(' ')
				speak("you have "+ temp[1] + " from "+time[0] + " to "+time[len(time)-1])

def tellmodule(day,hour):
	speak("On " + day +" At "+ hour +" you have " + getModulebyTime(day,hour))