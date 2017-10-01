import openpyxl as xl
import os
shldspath = {'schdls/cpi2.xlsx'}

def getsch():
	return shldspath
def getsheet(path,season,group):
	dir = os.listdir("schdls")
	if season in dir:
		wp = xl.load_workbook(path)
		sheets = wp.get_sheet_names()
		if group in sheets:
			return xl.get_sheet_by_name(group)

def getTds(day,sheet):
	days = list()
	modules = list()
	TDs = list()
	for i in range(2,sheet.max_column+1):
		days.append(sheet.cell(row=1,column=i).value)
	for i in days:
		if i == day:
			for j in range(2,sheet.max_rows+1):
					modules.append([ws.cell(row=j,column=1).value,ws.cell(row=j,column=days.index('Sunday')+2).value])
	for i in modules:
		s = i[1].split(' ')
		if s[0]="TD":
			TDs.append(i)
	return TDs

def getCourses(day,sheet):
	days = getdays(sheet)
	modules = list()
	cours = list()

	for i in days:
		if i == day:
			for j in range(2,sheet.max_rows+1):
					modules.append([ws.cell(row=j,column=1).value,ws.cell(row=j,column=days.index('Sunday')+2).value])
	for i in modules:
		s = i[1].split(' ')
		if s[0]="Cours":
			cours.append(i)
	return cours

def getAll(day,sheet):
	days = list()
	modules = list()
	cours = list()
	for i in range(2,sheet.max_column+1):
		days.append(sheet.cell(row=1,column=i).value)
	for i in days:
		if i == day:
			for j in range(2,sheet.max_rows+1):
					modules.append([ws.cell(row=j,column=1).value,ws.cell(row=j,column=days.index('Sunday')+2).value])
	return modules
def getTimeofmodule(module,sheet):
	days = getdays(sheet)
	for i in days:
		tds = getTds(i,sheet)
		for j in tds:
			if td in j[1]:
				return j[0]

def getdays(sheet):
	days = list()
	for i in range(2,sheet.max_column+1):
		days.append(sheet.cell(row=1,column=i).value)
	return days